import os

from Bio import SeqIO, Seq
from Bio.Seq import complement
from openpyxl import Workbook


ids = ['F3', 'F2', 'F1c', 'B1c', 'B2', 'B3']
colors = ['#FD0100', '#F76915', '#E6BC05', '#A0D636', '#2FA236', '#2E8BC0']
ids_loop = ['LF', 'BF']
colors_loop = ['#A9A9A9', '#708090']
id_probe = ['HP']
color_probe = ['#C5007F']


class Results:
    """
    Class for saving results
    """
    primer_sets = None
    cur_primer_set_ind = 0
    current_record = None


def parse_sequence_file(file_path: str) -> list:
    """
    Parse sequnce file (Fasta or Genbank)

    :param file_path: path to file

    :return: list with parsed records from file
    """
    return [record for record in SeqIO.parse(file_path, 'fasta')]


def get_filename(file_path: str) -> str:
    """
    Get filename without path

    :param file_path: path to file

    :return: filename
    """
    return os.path.basename(file_path)


def get_record_ids(records: list) -> list:
    """
    Get ids from records (work with Fasta files)

    :param records: list of records

    :return: list of ids of records 
    """
    return [record.id for record in records]


def get_record_params(record: Seq) -> str:
    """
    Parsing record

    :param record: record from parsed file

    :return: info about record
    """
    return 'ID: {}\nDescription: {}\nLength: {} bp'.format(
        record.id,
        record.description,
        len(record.seq)   
    )


def get_colors_and_ids(primer_set: list[list]) -> list[list, list]:
    """
    Написать
    """
    if len(primer_set) == 6:
        return ids, colors
    elif len(primer_set) == 7:
        return (
            ids[:3] + id_probe + ids[3:],
            colors[:3] + color_probe + colors[3:],
        )
    elif len(primer_set) == 8:
        return (
            ids[:2] + [ids_loop[0]] + ids[2:4] + [ids_loop[1]] + ids[4:],
            colors[:2] + [colors_loop[0]] + colors[2:4] + [colors_loop[1]] + colors[4:]
        )
    elif len(primer_set) == 9:
        return (
            ids[:2] + [ids_loop[0]] + [ids[2]] + id_probe + [ids[3]] + [ids_loop[1]] + ids[4:],
            colors[:2] + [colors_loop[0]] + [colors[2]] + color_probe + [colors[3]] + [colors_loop[1]] + colors[4:]
        )


def primer_sets_table(primer_set: list[list]) -> str:
    """
    Function to build table with primer set for UI

    :param primer_set: set with primers

    :return: html table with primers    
    """
    cur_ids, cur_colors = get_colors_and_ids(primer_set)

    primer_set_html = ''

    # Write primer info to row
    for index, primer in enumerate(primer_set):
        primer_set_html += (
            '<tr>'
                f'<td style="text-align: center;background-color: {cur_colors[index]};">{cur_ids[index]}</td>'
                f'<td width="{len(primer[0]) * 12}" style="font-family: Courier,Courier New,Monaco,monospace;">{str(primer[0]).upper()}</td>'
                f'<td style="text-align: center;">{primer[1][2]}</td>'
                f'<td style="text-align: center;">{primer[1][2] + primer[1][3] - 1}</td>'
                f'<td style="text-align: center;">{primer[1][3]}</td>'
                f'<td style="text-align: center;">{primer[1][0]}</td>'
                f'<td style="text-align: center;">{primer[1][1]}</td>'
            '</tr>'
        )   

    # Build FIP and BIP
    fip = primer_set[2][0] + ' ' + primer_set[1][0]
    bip = primer_set[4][0] + ' ' + primer_set[3][0]

    # Add BIP anf FIP to table
    primer_set_html += (
        '<tr>'
            '<td style="text-align: center;">FIP</td>'
            f'<td colspan="6" style="font-family: Courier,Courier New,Monaco,monospace;">{fip}</td>'
        '</tr>'
        '<tr>'
            '<td style="text-align: center;">BIP</td>'
            f'<td colspan="6" style="font-family: Courier,Courier New,Monaco,monospace;">{bip}</td>'
        '</tr>'
    )

    # Insert all rows to table
    primer_set_table = f"""<!DOCTYPE html>
                            <html lang="en">
                            <head>
                                <meta charset="UTF-8">
                            </head>
                            <body>
                                <table cellpadding="5" width="100%" border="1" align="left" style="font-family: Helvetica; font-size: 11pt;">
                                    <tr>
                                        <td style="text-align: center;"></td>
                                        <td height="15" style="text-align: center; ">Primer</td>
                                        <td style="text-align: center;">5'-pos</td>
                                        <td style="text-align: center;">3'-pos</td>
                                        <td style="text-align: center;">Length</td>
                                        <td style="text-align: center;">%GC</td>
                                        <td style="text-align: center;">Tm</td>
                                    </tr>
                                {primer_set_html}
                                </table>
                            </body>
                            </html>"""
    
    return primer_set_table


def get_sequences(primer_set: list[list]) -> str:    
    """
    Function to build sequence with primers in search window

    :param primer_set: set with primers

    :return: html window with sequences and primers
    """
    _, cur_colors = get_colors_and_ids(primer_set)
    # # Sort primers set by index for correct insert
    # primer_set.sort(key=lambda param: param[1][2])

    # Start/end indices
    inds = [
        primer_set[0][1][2] - primer_set[0][1][2] % 10 - 10,
        primer_set[-1][1][2] - primer_set[-1][1][2] % 10 + 31,
    ]

    cur_nucl_ind = 0
    cur_primer_ind = 0
    main_seq = ''
    compl_seq = ''
    indices = ''
    primers = ''

    # Sequence
    seq = str(Results.current_record.seq)

    # Fill table
    for i in range(inds[0], inds[1], 10):
        main_seq += f"""
                    <td><p style="font-family: Courier,Courier New,Monaco,monospace; font-style: normal; font-size: 12pt;">{seq[i:i + 10]}</p></td>
                """
        compl_seq += f"""
                    <td><p style="font-family: Courier,Courier New,Monaco,monospace; font-style: normal; font-size: 12pt;">{str(complement(seq[i:i + 10]))}</p></td>
                """
        indices += f"""
                    <td><p style="font-family: Courier,Courier New,Monaco,monospace; font-style: normal; font-size: 12pt;">{str(i)}</p></td>
                """
        
        # Fill colored primers in sequence
        if cur_primer_ind != len(primer_set):
            cur_nucl_string = ''
            space_counts = 0
            html_td = ''

            for j in range(10):
                if cur_nucl_ind == primer_set[cur_primer_ind][1][3]:
                    html_td += f"""<span style="background-color: {cur_colors[cur_primer_ind]};">{cur_nucl_string}</span>"""

                    cur_nucl_ind = 0
                    cur_primer_ind += 1

                    if cur_primer_ind == len(primer_set):
                        break

                    continue
                
                if i + j == primer_set[cur_primer_ind][1][2]:
                    space_counts = j - len(cur_nucl_string)
                    cur_nucl_string = ''
                    
                    cur_nucl_string += primer_set[cur_primer_ind][0][cur_nucl_ind]
                    cur_nucl_ind += 1

                elif i + j > primer_set[cur_primer_ind][1][2]:
                    cur_nucl_string += primer_set[cur_primer_ind][0][cur_nucl_ind]
                    cur_nucl_ind += 1
            
            if cur_nucl_string:
                if space_counts:
                    html_td += f"""<span>{'&nbsp;' * space_counts}</span>"""

                if cur_nucl_ind > 0:
                    html_td += f"""<span style="background-color: {cur_colors[cur_primer_ind]};">{cur_nucl_string}</span>"""

                primers += f"""<td><p style="font-family: Courier,Courier New,Monaco,monospace; font-style: normal; font-size: 12pt;">{html_td}</p></td>"""            
            else:
                primers += f"""<td></td>""" 

    # Insert table in html
    sequences_html = f"""<!DOCTYPE html>
                            <html lang="en">
                            <head>
                                <meta charset="UTF-8">
                            </head>
                            <body>
                                <table cel>
                                    <tr align="left">
                                        <td><p style="font-family: Helvetica; font-size: 11pt;">Target DNA</p></td>
                                        {main_seq}
                                    </tr>
                                    <tr align="left">
                                        <td style="padding-right: 20px; font-family: Helvetica; font-size: 11pt;">Complement DNA</td>
                                        {compl_seq}
                                    </tr>
                                    <tr>
                                        <td></td>
                                        {indices}
                                    </tr>
                                    <tr>
                                        <td></td>
                                        {primers}
                                    </tr>
                                </table>
                            </body>
                        </html>"""

    return sequences_html


def save_to_excel(primers_sets: list[list], file_path: str) -> bool:
    """
    Save results to excel

    :param primers_sets: sets of primers which should be saved
    :param file_path: path to Excel file

    :return: True, if file was created and saved, else False
    """
    try:
        ids = ['F3', 'F2', 'F1c', 'B1c', 'B2', 'B3']
        cols_names = ['Primer', '%GC', 'Tm', "5'-pos", "3'-pos", 'Length']

        wb = Workbook()
        sheet = wb.active
        sheet.column_dimensions['B'].width = 35
        
        sheet.title = 'Designed Primers'

        for count, primer_set in enumerate(primers_sets):
            cur_count_set = 10 * count + 1

            for ind in range(6):
                sheet.cell(row=cur_count_set, column=ind + 2).value = cols_names[ind]
                sheet.cell(row=cur_count_set + ind + 1, column=1).value = ids[ind]
                
                primer = primer_set[ind]
    
                params = [primer[0]] + primer[1][:2] + \
                    [primer[1][2], primer[1][2] + primer[1][3], primer[1][3]]
                    
                for param_ind, primer_param in enumerate(params):
                    sheet.cell(row=cur_count_set + ind + 1, column=param_ind + 2).value = primer_param

        wb.save(file_path)

    except Exception as e:
        return False

